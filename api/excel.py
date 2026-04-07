from http.server import BaseHTTPRequestHandler
import json, os
from io import BytesIO
from datetime import datetime
import openpyxl

DOW = ["일","월","화","수","목","금","토"]
PY_TO_KR = [1,2,3,4,5,6,0]

def fmt_date(ds):
    d = datetime.strptime(ds, '%Y-%m-%d')
    return f"{d.month}월 {d.day}일", DOW[PY_TO_KR[d.weekday()]]

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length)
            data = json.loads(body)

tpl = '/var/task/public/alpensia_template.xlsx'            if not os.path.exists(tpl):
                tpl = '/var/task/public/alpensia_template.xlsx'

            wb = openpyxl.load_workbook(tpl)
            ws = wb.active

            today = datetime.now()
            ws['G4'] = f"{today.year}년"
            ws['H4'] = f"{today.month}월"
            ws['J4'] = f"{today.day}일"
            ws['B6'] = data.get('repName', '')
            ws['G6'] = data.get('phone', '')

            for i, g in enumerate(data.get('golfRounds', [])[:4]):
                row = 11 + i
                date_str, dow = fmt_date(g['date'])
                ws[f'A{row}'] = date_str
                ws[f'B{row}'] = dow
                ws[f'C{row}'] = g.get('course', '')
                ws[f'D{row}'] = g.get('teams', 1)
                ws[f'E{row}'] = g.get('ppl', 4)
                ws[f'F{row}'] = g.get('tee', '')
                ws[f'H{row}'] = g.get('gf', 0)
                if g.get('bf', 0) > 0:
                    ws[f'J{row}'] = g.get('bf', 0)

            for i, rm in enumerate(data.get('rooms', [])[:3]):
                row = 20 + i
                date_str, dow = fmt_date(rm['date'])
                ws[f'A{row}'] = date_str
                ws[f'B{row}'] = dow
                ws[f'C{row}'] = rm.get('type', '')
                ws[f'D{row}'] = rm.get('spec', '')
                ws[f'F{row}'] = rm.get('cnt', 1)
                ws[f'G{row}'] = rm.get('rate', 0)

            if data.get('memo'):
                ws['B27'] = data.get('memo', '')

            out = BytesIO()
            wb.save(out)
            out.seek(0)
            xlsx = out.read()

            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="알펜시아예약신청서.xlsx"')
            self.send_header('Content-Length', str(len(xlsx)))
            self.end_headers()
            self.wfile.write(xlsx)

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())

    def log_message(self, fmt, *args):
        pass
